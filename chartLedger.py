#imports
from re import search
from openpyxl import load_workbook, Workbook

#function fillChart that will fill de chart file with values from ledger
def fillChart(chart, ledger):

    """
    fillChart Function
    This function needs two parameters, both required and both must be a Excel ou similar file
    ------------------------------------------
    First Parameter
    ---> chart 
        It must contain a single column with the numbers of the charts it will search
    
    Second Parameter
    ---> ledger
        It mus contain two columns:
            A Column - Number of charts
            B Column - Value in the chart
    ------------------------------------------
    Operation
    ------------------------------------------
    ---> The Search
        The function will take the input files and assign both in a variable for each one.
        Next, the function will get a value from "sheetchart" variable and searches this value in the first column  from "sheetLedger" variable.
        If the values are equals, it get a value from respective row, but from second column and add in the "valueBColumn" variable. 
        This variable will be assign in the output file. 
    ------------------------------------------
    ---> The combination of Values
        The second part of the code will combine values from the same branch of the tree.
        First, the code looks for cells where the values are equal to zero (if different, it skips to the next iteration).
        Then store that value and string length in separate variables.
        The value will be the search key for sub-values, and the length will be the limiter to not get values from different branches.  
    """

    #created a .XLSX file to fill with new datas
    out_chart_of_accounts = Workbook()
    out_plan1 = out_chart_of_accounts.active
    out_plan1.title = 'Ex_Chart_Accounts'

    #create the header from output file
    out_plan1.cell(row=1, column=1, value='account')
    out_plan1.cell(row=1, column=2, value='value')

    sheetChart = chart.active #activated the sheet from origin file chart_of_accounts and assing to the sheet1 variable.
    sheetLedger = ledger.active
    maxrowChart = sheetChart.max_row #take the last row from sheet Chart
    maxrowLedger = sheetLedger.max_row #take the last row from sheet Ledger

    #first loop. Enter in the chart_of_accounts file to get a value to search
    for i in range(2, maxrowChart+1): 
        valueBColumn = float(0.0)
        searchValue = sheetChart.cell(row=i, column=1).value #value that will be searched in the "ledger" file
        
        if searchValue == None: #Jump the remaining loop if get a empty cell (generally is in the end of file)
            continue

        #Second loop. Enter in the general_ledger file to search and sum values from var "searchValue"
        for j in range(2, maxrowLedger+1):
            valueCh = sheetLedger.cell(row=j, column=1).value #get chart name
            valueLe = sheetLedger.cell(row=j, column=2).value #get chart value
            try:
                valueLeFl = round(float(valueLe), 2) #convert str to float

                #if the values are equal, increment in the var valueBColumn
                if valueCh == searchValue:
                    valueBColumn += valueLeFl
            except:
                #Probable error when converting to float 
                continue            

        try:    #write values from columns A and B in the output file, with a ERROR test
            out_plan1.cell(row=i, column=1, value=searchValue)
            out_plan1.cell(row=i, column=2, value=valueBColumn)
        except:
            print('Error! Impossible save the file!')

    #Second part! Combination of values
    #-------------------------------------------------------------

    max_rowOut = out_plan1.max_row #take the last row from sheet out_plan1

    #first loop. It get a first value equal zero, and search subvalues to add.
    for i in range(2, max_rowOut+1):
        
        valueOutV = out_plan1.cell(row=i, column=2).value

        if valueOutV != 0: #if the value from B column not be zero, it jump the loop
            continue
        else:
            valueOutC = out_plan1.cell(row=i, column=1).value #value that will be used to get subvalues
            newSum = 0.0
            lenGetValue = len(valueOutC) #get a length from origin value. It will be a paramenter for limit of subvalues

            #Second loop. This will search for subvalues
            for j in range(2, max_rowOut+1):
                tempC = out_plan1.cell(row=j, column=1).value

                try:
                    tempV = round(float(out_plan1.cell(row=j, column=2).value), 2)

                    #if the subvalue equals search value, this will be add to var 'newSum'
                    if valueOutC == tempC[:lenGetValue]:
                        newSum += tempV
                except:
                    #Probable error when converting to float 
                    continue

            #write the newSum value in the output file
            out_plan1.cell(row=i, column=2, value=newSum)

    #save the output file in the "output" diretory and close it
    try:
        out_chart_of_accounts.save('output/out_chart_of_accounts.xlsx')
        out_chart_of_accounts.close()
    except:
        print('Error! Unable to save file. Check write permission for the folder!')
    #RETURN
    #None file will be returned. The new file will be saved in the "output" diretory

#load files from input diretory

try:
    chart_of_accounts = load_workbook('input/chart_of_accounts.xlsx')
    general_ledger = load_workbook('input/general_ledger.xlsx')

    fillChart(chart_of_accounts, general_ledger)
except:
    print('Error! Unable to load files!')
